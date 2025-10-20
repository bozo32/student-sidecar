#!/usr/bin/env python3
"""
process_submissions.py — walk subdirectories and capture a 4-row head from each CSV/XLS(X).
Optionally discover URLs embedded in tables, fetch them, and save into each group's folder.

Usage:
  python process_submissions.py /path/to/root [--fetch-urls] [--urls-subdir urls] [--url-timeout 20]

Outputs:
  - Console summary (file path, type, sheet, inferred delimiter, columns)
  - Per-file/per-sheet previews saved under ./previews/ as CSV
  - (optional) For each table, fetched URLs saved under <group>/<urls-subdir>/ with a per-group _manifest.csv
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional, Tuple, List, Dict

import requests
import pandas as pd

# ---------- Configuration ----------

PREVIEW_DIR = Path("previews")
PREVIEW_DIR.mkdir(parents=True, exist_ok=True)

# canonical headers we’ll normalize to
CANON = ["quote_from_report", "file_name", "quote_from_source"]

# simple URL recognizer (tolerant)
URL_RE = re.compile(r"""(?ix)
\b(https?://[^\s<>'"]+)
""")

# columns that are likely to contain URLs
LIKELY_URL_COLNAMES = {
    "url", "link", "source_url", "web", "website", "doi", "pdf_url"
}

# (Optional) If you ever want to only fetch “document-like” URLs, you can use this list.
DOC_EXTS = (".pdf", ".doc", ".docx", ".rtf", ".txt")


@dataclass
class TableSpec:
    group_dir: Path
    table_path: Path
    is_excel: bool
    sheet_name: Optional[str] = None


def find_tables(root: Path) -> Iterable[TableSpec]:
    """
    Yield TableSpec for each CSV or Excel file under root/<group>/*.
    Preference order for a group: if both CSV and XLSX exist with same stem,
    the Excel version will be *preferred* later (we’ll still preview both).
    """
    for group in sorted(p for p in root.iterdir() if p.is_dir()):
        csvs = sorted(group.glob("*.csv"))
        xls = sorted(group.glob("*.xls")) + sorted(group.glob("*.xlsx"))
        for p in csvs:
            yield TableSpec(group, p, is_excel=False)
        for p in xls:
            yield TableSpec(group, p, is_excel=True)


def normalize_header_names(cols: List[str]) -> List[str]:
    """
    Map common variants to canonical headers: quote_from_report, file_name, quote_from_source.
    Leaves everything else as-is. We’ll collapse to first 3 logical columns later if needed.
    """
    out = []
    for c in cols:
        if c is None:
            out.append(c)
            continue
        k = str(c).strip().lower()
        k = re.sub(r"\s+", " ", k)
        k = k.replace("’", "'").replace("“", '"').replace("”", '"')

        if k in {"quote from report", "text in report", "quote in our report", "assignment 1 citations"}:
            out.append("quote_from_report")
        elif k in {"file name", "file name ", "filename", "source", "source file", "file", "file_name"}:
            out.append("file_name")
        elif k in {"quote from source", "used text from source", "text from source", "quote from resource", "qoute from source"}:
            out.append("quote_from_source")
        else:
            out.append(k)
    return out


def try_read_csv(path: Path) -> Tuple[pd.DataFrame, Optional[str]]:
    """
    Robust CSV reader: sniff delimiter, try multiple encodings and engines,
    tolerate bad lines. Returns (df, delimiter_guess)
    """
    # Attempt basic sniff (best-effort)
    delim = None
    try:
        with path.open("r", encoding="utf-8", errors="ignore") as f:
            sample = f.read(4096)
        sniffer = csv.Sniffer()
        dialect = sniffer.sniff(sample)
        delim = dialect.delimiter
    except Exception:
        delim = None

    encodings = ["utf-8", "utf-8-sig", "cp1252", "latin-1"]
    engines = ["c", "python"]

    last_err = None
    for enc in encodings:
        for eng in engines:
            try:
                df = pd.read_csv(
                    path,
                    dtype=str,
                    on_bad_lines="skip",
                    engine=eng,
                    encoding=enc,
                    sep=(delim if delim else None),  # prefer the sniffer if we got one
                    encoding_errors="ignore",
                )
                return df, delim
            except TypeError:
                try:
                    df = pd.read_csv(
                        path,
                        dtype=str,
                        on_bad_lines="skip",
                        engine=eng,
                        encoding=enc,
                        sep=(delim if delim else None),
                    )
                    return df, delim
                except Exception as e:
                    last_err = e
            except Exception as e:
                last_err = e

    raise RuntimeError(f"[READ-ERROR] {path}: {last_err}")


def read_excel_sheets(path: Path) -> Dict[str, pd.DataFrame]:
    """Read all sheets from Excel (xlsx/xls) as str DF."""
    try:
        xls = pd.ExcelFile(path)
        out: Dict[str, pd.DataFrame] = {}
        for sheet in xls.sheet_names:
            df = pd.read_excel(path, sheet_name=sheet, dtype=str, header=0)
            out[sheet] = df
        return out
    except Exception as e:
        raise RuntimeError(f"[READ-ERROR] {path}: {e}")


def get_excel_hyperlinks(path: Path) -> Dict[str, List[str]]:
    """
    Return a mapping of sheet_name -> list of hyperlink targets found via openpyxl.
    Only works for .xlsx; silently returns {} for .xls or if openpyxl is unavailable.
    """
    links_by_sheet: Dict[str, List[str]] = {}
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        return links_by_sheet

    if path.suffix.lower() != ".xlsx":
        return links_by_sheet

    try:
        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
        for ws in wb.worksheets:
            found = []
            for row in ws.iter_rows():
                for cell in row:
                    try:
                        hl = getattr(cell, "hyperlink", None)
                        if hl:
                            target = getattr(hl, "target", None) or getattr(hl, "ref", None)
                            if target and isinstance(target, str):
                                found.append(target)
                    except Exception:
                        continue
            if found:
                links_by_sheet[ws.title] = found
    except Exception:
        return links_by_sheet
    return links_by_sheet


def collapse_to_first_three(df: pd.DataFrame) -> pd.DataFrame:
    """
    If there are more than three “interesting” columns, try to collapse to the first 3 that look like
    our canonical trio. Otherwise keep/pad the first 3 visible columns.
    Always returns a DataFrame with up to 3 columns (pads with blanks if fewer).
    """
    if df is None or df.shape[1] == 0:
        return pd.DataFrame({"_col0": []})

    cols = normalize_header_names(list(df.columns))
    df = df.copy()
    df.columns = cols

    has_canon = [c for c in CANON if c in df.columns]
    if len(has_canon) == 3:
        out = df[CANON].copy()
    else:
        picks: List[str] = []
        for target in CANON:
            for c in df.columns:
                if c == target and c not in picks:
                    picks.append(c)
                    break
        for c in df.columns:
            if len(picks) >= 3:
                break
            if c not in picks:
                picks.append(c)
        out = df[picks[:3]].copy()

    while out.shape[1] < 3:
        out[f"_pad{out.shape[1]}"] = ""
    out = out.iloc[:, :3]
    return out


def extract_urls_from_df(df: pd.DataFrame) -> List[str]:
    urls = set()
    for col in df.columns:
        if str(col).strip().lower() in LIKELY_URL_COLNAMES:
            for v in df[col].astype(str).fillna(""):
                for m in URL_RE.findall(v):
                    urls.add(m)
        else:
            for v in df[col].astype(str).fillna(""):
                for m in URL_RE.findall(v):
                    urls.add(m)
    return sorted(urls)


def sanitize_filename(name: str) -> str:
    # remove pathy bits and illegal characters
    name = name.split("/")[-1]
    name = re.sub(r'[\\:*?"<>|]+', "_", name).strip()
    return name or "download.bin"


def fetch_url(url: str, dest: Path, timeout: int = 20) -> Optional[Path]:
    """
    Download URL to dest. Prefer filename from Content-Disposition; fall back to URL tail.
    """
    try:
        r = requests.get(url, timeout=timeout, headers={"User-Agent": "student-sidecar/1.0"})
        r.raise_for_status()

        fname = None
        cd = r.headers.get("Content-Disposition", "")
        # try RFC 5987 or simple filename=
        m = re.search(r'filename\*?=(?:UTF-8\'\')?"?([^";]+)"?', cd, re.I)
        if m:
            fname = sanitize_filename(m.group(1))
        if not fname:
            tail = url.split("/")[-1].split("?")[0]
            fname = sanitize_filename(tail if "." in tail else "download.bin")

        out_path = dest / fname
        with out_path.open("wb") as f:
            f.write(r.content)
        return out_path
    except Exception:
        return None


def ensure_manifest(dirpath: Path) -> Path:
    dirpath.mkdir(parents=True, exist_ok=True)
    m = dirpath / "_manifest.csv"
    if not m.exists():
        with m.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=["timestamp", "url", "saved_as"])
            w.writeheader()
    return m


def append_manifest_row(manifest_path: Path, row: Dict[str, str]) -> None:
    with manifest_path.open("a", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["timestamp", "url", "saved_as"])
        w.writerow(row)


def _download_urls(urls: List[str], group_dir: Path, urls_subdir: str, url_timeout: int) -> None:
    """Helper: download URLs into <group>/<urls_subdir>/, skipping duplicates, logging to _manifest.csv."""
    if not urls:
        return
    url_dir = group_dir / urls_subdir
    manifest = ensure_manifest(url_dir)

    # set of already-existing filenames (for cheap dedupe)
    existing = {p.name for p in url_dir.glob("*") if p.is_file()}

    for u in urls:
        if not u.lower().startswith("http"):
            continue  # ignore local/file paths
        # OPTIONAL: limit to doc-like URLs only
        # if not any(u.lower().split("?", 1)[0].endswith(ext) for ext in DOC_EXTS):
        #     continue

        # Guess target name from URL to avoid needless re-fetch
        guess = sanitize_filename(u.split("/")[-1].split("?")[0] or "download.bin")
        if guess in existing:
            append_manifest_row(manifest, {
                "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
                "url": u,
                "saved_as": str(url_dir / guess)
            })
            continue

        saved = fetch_url(u, url_dir, timeout=url_timeout)
        append_manifest_row(manifest, {
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
            "url": u,
            "saved_as": str(saved) if saved else ""
        })


def preview_and_optionally_fetch(spec: TableSpec, fetch_urls: bool, urls_subdir: str, url_timeout: int) -> None:
    """
    Read table, print preview to console, write 4-row head to previews/.
    If fetch_urls=True, also download any URLs found into <group>/<urls_subdir>/ and
    append to a per-group _manifest.csv.
    """
    rel_group = spec.group_dir.name
    if spec.is_excel:
        sheets = read_excel_sheets(spec.table_path)
        print(f"[XLS/XLSX] {spec.table_path}  sheets={list(sheets.keys())}")

        xl_hlinks = get_excel_hyperlinks(spec.table_path)

        for sheet, df in sheets.items():
            if df is None or df.shape[1] == 0:
                print(f"   ↳ sheet='{sheet}'  (empty)  SKIPPED")
                continue
            df = df.astype(str)
            df_norm = collapse_to_first_three(df).copy()
            if df_norm.shape[1] < 3:
                while df_norm.shape[1] < 3:
                    df_norm[f"_pad{df_norm.shape[1]}"] = ""
                df_norm = df_norm.iloc[:, :3]
            df_norm.columns = CANON
            head = df_norm.head(4)
            out_name = f"{rel_group}_{spec.table_path.stem}_{sheet}_head.csv"
            head.to_csv(PREVIEW_DIR / out_name, index=False)
            print(f"   ↳ sheet='{sheet}'  columns={list(df_norm.columns)}  rows_shown={len(head)}")

            if fetch_urls:
                urls_text = set(extract_urls_from_df(df))
                urls_link = set(xl_hlinks.get(sheet, []))
                urls = sorted(urls_text.union(urls_link))
                # Filter to http(s) only (ignore local files)
                urls = [u for u in urls if u.lower().startswith("http")]
                _download_urls(urls, spec.group_dir, urls_subdir, url_timeout)

    else:
        df, delim = try_read_csv(spec.table_path)
        df = df.astype(str)
        df_norm = collapse_to_first_three(df).copy()
        if df_norm.shape[1] < 3:
            while df_norm.shape[1] < 3:
                df_norm[f"_pad{df_norm.shape[1]}"] = ""
            df_norm = df_norm.iloc[:, :3]
        df_norm.columns = CANON
        head = df_norm.head(4)
        out_name = f"{rel_group}_{spec.table_path.stem}_head.csv"
        head.to_csv(PREVIEW_DIR / out_name, index=False)
        print(f"[CSV] {spec.table_path}\n      delimiter={repr(delim)}  columns={list(df_norm.columns)}  rows_shown={len(head)}")

        if fetch_urls:
            urls = extract_urls_from_df(df)
            urls = [u for u in urls if u.lower().startswith("http")]
            _download_urls(urls, spec.group_dir, urls_subdir, url_timeout)


def main():
    ap = argparse.ArgumentParser(description="Preview student tables and optionally fetch embedded URLs.")
    ap.add_argument("root", help="Root folder containing group subfolders")
    ap.add_argument("--fetch-urls", action="store_true", help="Fetch any URLs found in tables")
    ap.add_argument("--urls-subdir", default="urls", help="Per-group subfolder to save fetched URLs (default: urls)")
    ap.add_argument("--url-timeout", type=int, default=20, help="HTTP timeout seconds per URL (default: 20)")
    args = ap.parse_args()

    root = Path(args.root).expanduser().resolve()
    if not root.exists():
        print(f"[ERROR] Root not found: {root}", file=sys.stderr)
        sys.exit(2)

    any_found = False
    for spec in find_tables(root):
        any_found = True
        try:
            preview_and_optionally_fetch(spec, fetch_urls=args.fetch_urls, urls_subdir=args.urls_subdir, url_timeout=args.url_timeout)
        except Exception as e:
            print(str(e), file=sys.stderr)

    if not any_found:
        print(f"[INFO] No CSV/XLS workbooks found under {root}")


if __name__ == "__main__":
    main()
